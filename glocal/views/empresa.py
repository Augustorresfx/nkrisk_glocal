from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponseRedirect
from django.views.generic import View
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.shortcuts import redirect
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.shortcuts import get_list_or_404

# Importe Modelos
from ..models import Pais, Broker, PendingChange, Matriz, Contacto, Empresa

# REPORTE
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from django.http import HttpResponse

# ADMINISTRACION
@method_decorator(login_required, name='dispatch')
class EmpresaView(View):
    def get(self, request, *args, **kwargs):
        
        # Obtener filtros desde los parámetros GET
        nombre_filtro = request.GET.get('nombre', '').strip()
        pais_filtro = request.GET.get('pais', '').strip()

        # Filtrar con base en los filtros
        empresas = Empresa.objects.all()

        if nombre_filtro:
            empresas = empresas.filter(nombre__icontains=nombre_filtro)

        if pais_filtro:
            empresas = empresas.filter(pais__id=pais_filtro)


        exportar = request.GET.get('exportar', None)

        # Si se solicita exportar, generar el archivo Excel
        if exportar:
            return self.generar_excel_empresas(empresas)

        # Paginación
        empresas_paginados = Paginator(empresas, 30)
        page_number = request.GET.get("page")
        filter_pages = empresas_paginados.get_page(page_number)

        # Obtener la lista de países
        paises = Pais.objects.all()

        matrices = Matriz.objects.all()

        contactos = Contacto.objects.all()

        context = {
            'empresas': empresas,
            'pages': filter_pages,
            'paises': paises,
            'matrices': matrices,
            'contactos': contactos,
            'filtros': {
                'nombre': nombre_filtro,
                'pais': pais_filtro,
            }
        }
        return render(request, 'administracion/empresas_admin.html', context)
    def post(self, request, *args, **kwargs):
        # Obtener datos del formulario
        nombre = request.POST.get('nuevo_nombre')
        ruc_nit = request.POST.get('nuevo_ruc_nit')
        matriz_id = request.POST.get('nuevo_matriz')
        pais_id = request.POST.get('nuevo_pais')
        activo = request.POST.get('nuevo_activo')
        contactos_ids = request.POST.getlist('nuevo_contacto') 

        if activo == "on":
            activo = True
        else:
            activo = False
        # Obtener objetos relacionados
        pais = get_object_or_404(Pais, id=pais_id)
        matriz = get_object_or_404(Matriz, id=matriz_id)
        contactos = get_list_or_404(Contacto, id__in=contactos_ids)
        user = request.user
        # Crear una solicitud de creación pendiente
        changes = {
            'nombre': {'new': nombre},
            'ruc_nit': {'new': ruc_nit},
            'pais_id': {'new': pais.id},
            'matriz_id': {'new': matriz.id},
            'activo': {'new': activo},
            'contactos': {'new': [c.id for c in contactos]},  # IDs de los contactos seleccionados
        }

        if user.is_superuser:
            try:
                # Crear el objeto directamente si el usuario es superusuario
                empresa = Empresa.objects.create(
                    nombre=nombre,
                    ruc_nit=ruc_nit,
                    pais=pais,
                    matriz=matriz,
                    activo=activo,
                    modified_by = user,
                )
                empresa.contactos.set(contactos)  # Relación muchos a muchos
                messages.success(request, 'Se creó un nuevo elemento de forma exitosas.')
            except Exception as e:
                messages.error(request, f'Error: No se pudo crear el elemento. Detalles: {str(e)}')
 
        else:
            try:
                PendingChange.objects.create(
                    model_name='empresa',
                    object_id=None,  # No existe aún
                    changes=changes,
                    submitted_by=user,
                    action_type='create',
                )
                messages.success(request, 'Solicitud de creación enviada para aprobación.')
            except Exception as e:
                messages.error(request, f'Error: No se pudo enviar la solicitud. Detalles: {str(e)}')

        return HttpResponseRedirect(request.path_info)
    
    def generar_excel_empresas(self, empresas):
        # Crear un nuevo archivo Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "empresas"

        # Estilos
        font_header = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        relleno_header = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        bordes = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )

        # Encabezados
        encabezados = ["Nombre", "Matriz", "País", "RUC / NIT", "Activo", "Contactos"]
        for col_num, header in enumerate(encabezados, start=1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = font_header
            cell.fill = relleno_header
            cell.border = bordes

        # Datos
        for row_num, empresa in enumerate(empresas, start=2):
            # Columna 1: Nombre
            cell_nombre = sheet.cell(row=row_num, column=1, value=empresa.nombre)
            cell_nombre.border = bordes

            # Columna 2: Matriz
            cell_matriz = sheet.cell(row=row_num, column=2, value=empresa.matriz.nombre)
            cell_matriz.border = bordes

            # Columna 3: País
            cell_pais = sheet.cell(row=row_num, column=3, value=empresa.pais.nombre)
            cell_pais.border = bordes

            # Columna 4: RUC / NIT
            cell_ruc_nit = sheet.cell(row=row_num, column=4, value=empresa.ruc_nit)
            cell_ruc_nit.border = bordes

            # Columna 5: Activo
            cell_activo = sheet.cell(row=row_num, column=5, value="SI" if empresa.activo else "NO")
            cell_activo.border = bordes

            # Columna 6: Contactos
            # Obtener nombres de contactos separados por comas
            contactos_nombres = ", ".join(contacto.nombre for contacto in empresa.contactos.all())
            sheet.cell(row=row_num, column=6, value=contactos_nombres).border = bordes

        # Crear la respuesta HTTP
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="empresas.xlsx"'
        workbook.save(response)
        workbook.close()
        return response


@method_decorator(login_required, name='dispatch')
class EditarEmpresaView(View):
    def post(self, request, empresa_id):
        empresa = get_object_or_404(Empresa, id=empresa_id)
        nombre = request.POST.get('editar_nombre')
        pais_id = request.POST.get('editar_pais')
        ruc_nit = request.POST.get('editar_ruc_nit')
        matriz_id = request.POST.get('editar_matriz')
        activo = 'editar_activo' in request.POST
        contactos_ids = request.POST.getlist('editar_contacto') 
        
        # Obtener objetos relacionados
        pais = get_object_or_404(Pais, id=pais_id)
        matriz = get_object_or_404(Matriz, id=matriz_id)
        contactos = get_list_or_404(Contacto, id__in=contactos_ids)
        user = request.user

        # Si el usuario es superuser, aplicar los cambios directamente
        if user.is_superuser:
            if empresa.nombre != nombre:
                empresa.nombre = nombre
            if empresa.pais_id != int(pais_id):
                empresa.pais_id = pais_id
            if empresa.ruc_nit != ruc_nit:
                empresa.ruc_nit = ruc_nit
            if empresa.matriz_id != int(matriz_id):
                empresa.matriz_id = matriz_id
            if empresa.activo != activo:
                empresa.activo = activo

            empresa.contactos.set(contactos)  # Actualiza los contactos asociados
            empresa.modified_by = user  # Registrar quién realizó el cambio
            empresa.save(track_changes=False)  # Evitar registrar PendingChange
            messages.success(request, 'Los cambios se han aplicado directamente.')
            return redirect('empresas_admin')
        else:
            # Para usuarios no superuser, registrar los cambios pendientes
            changes = {}
            if empresa.nombre != nombre:
                changes['nombre'] = {'old': empresa.nombre, 'new': nombre}
                empresa.nombre = nombre

            if empresa.pais_id != int(pais_id):
                changes['pais_id'] = {'old': empresa.pais_id, 'new': int(pais_id)}
                empresa.pais_id = pais_id

            if empresa.ruc_nit != ruc_nit:
                changes['ruc_nit'] = {'old': empresa.ruc_nit, 'new': ruc_nit}
                empresa.ruc_nit = ruc_nit

            if empresa.matriz_id != int(matriz_id):
                changes['matriz_id'] = {'old': empresa.matriz_id, 'new': int(matriz_id)}
                empresa.matriz_id = matriz_id

            if empresa.activo != activo:
                changes['activo'] = {'old': empresa.activo, 'new': activo}
                empresa.activo = activo

            # Comparar contactos, actualizando la relación ManyToMany
            if set(empresa.contactos.all()) != set(contactos):
                changes['contactos'] = {
                    'old': [contacto.id for contacto in empresa.contactos.all()],
                    'new': [contacto.id for contacto in contactos]
                }

            if changes:
                try:
                    PendingChange.objects.create(
                        model_name='empresa',    
                        object_id=empresa.id,
                        changes=changes,
                        submitted_by=user,
                        action_type='edit'
                    )
                    messages.success(request, 'El cambio ha sido registrado para revisión.')
                except Exception as e:
                    messages.error(request, f'Error al registrar el cambio: {str(e)}')
            else:
                messages.info(request, 'No se detectaron cambios.')

        return redirect('empresas_admin')
    
@method_decorator(login_required, name='dispatch')
class EliminarEmpresaView(View):
    def post(self, request, empresa_id):
        empresa = get_object_or_404(Empresa, id=empresa_id)
        user = request.user

        if user.is_superuser:
            empresa.delete()
            messages.success(request, 'El elemento ha sido eliminado directamente.')
            return redirect('empresas_admin')

        changes = {
            'activo': {'old': empresa.activo, 'new': False},  # Simula desactivación
        }
        try:
            PendingChange.objects.create(
                model_name='empresa',
                object_id=empresa.id,
                changes=changes,
                submitted_by=user,
                action_type='delete'  # Indica que es una solicitud de eliminación
            )
            messages.success(request, 'El cambio de eliminación se ha registrado para revisión.')
        except Exception as e:
            messages.error(request, f'Error al registrar el cambio: {str(e)}')

        return redirect('empresas_admin')

# RESUMEN
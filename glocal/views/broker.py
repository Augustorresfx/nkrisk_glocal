from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponseRedirect
from django.views.generic import View
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.shortcuts import redirect
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.shortcuts import get_list_or_404

# Importe Modelos
from ..models import Pais, Broker, PendingChange, Matriz, Contacto

# REPORTE
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from django.http import HttpResponse

# ADMINISTRACION
@method_decorator(login_required, name='dispatch')
class BrokerView(View):
    def get(self, request, *args, **kwargs):
        
        # Obtener filtros desde los parámetros GET
        nombre_filtro = request.GET.get('nombre', '').strip()
        pais_filtro = request.GET.get('pais', '').strip()
        exportar = request.GET.get('exportar', None)
        # Filtrar los brokers con base en los filtros
        brokers = Broker.objects.all()

        if nombre_filtro:
            brokers = brokers.filter(nombre__icontains=nombre_filtro)

        if pais_filtro:
            brokers = brokers.filter(pais__id=pais_filtro)

        # Paginación
        brokers_paginados = Paginator(brokers, 30)
        page_number = request.GET.get("page")
        filter_pages = brokers_paginados.get_page(page_number)

        # Si se solicita exportar, generar el archivo Excel
        if exportar:
            return self.generar_excel_brokers(brokers)

        # Obtener la lista de países
        paises = Pais.objects.all()

        matrices = Matriz.objects.all()

        contactos = Contacto.objects.all()

        context = {
            'brokers': brokers,
            'pages': filter_pages,
            'paises': paises,
            'matrices': matrices,
            'contactos': contactos,
            'filtros': {
                'nombre': nombre_filtro,
                'pais': pais_filtro,
            }
        }
        return render(request, 'administracion/brokers_admin.html', context)
    def post(self, request, *args, **kwargs):
        # Obtener datos del formulario
        nombre = request.POST.get('nuevo_nombre')
        logo = request.FILES.get('nuevo_logo')
        domicilio = request.POST.get('nuevo_domicilio')
        web = request.POST.get('nuevo_web')
        matriz_id = request.POST.get('nuevo_matriz')
        pais_id = request.POST.get('nuevo_pais')
        activo = request.POST.get('nuevo_activo')
        contactos_ids = request.POST.getlist('nuevo_contacto')
        if activo == "on":
            activo = True
        else:
            activo = False
        # Obtener objetos relacionados
        pais = get_object_or_404(Pais, id=pais_id)
        matriz = get_object_or_404(Matriz, id=matriz_id)
        contactos = get_list_or_404(Contacto, id__in=contactos_ids)
        user = request.user
        # Crear una solicitud de creación pendiente
        changes = {
            'nombre': {'new': nombre},
            'logo': {'new': logo.name if logo else None},  # Guardar el nombre del archivo
            
            'domicilio_oficina': {'new': domicilio},
            'url_web': {'new': web},
            'pais_id': {'new': pais.id},
            'matriz_id': {'new': matriz.id},
            'activo': {'new': activo},
            'contactos': {'new': [c.id for c in contactos]},  # IDs de los contactos seleccionados
        }

        if user.is_superuser:
            try:
                # Crear el objeto Broker directamente si el usuario es superusuario
                broker = Broker.objects.create(
                    nombre=nombre,
                    logo=logo,
                    domicilio_oficina=domicilio,
                    url_web=web,
                    pais=pais,
                    matriz=matriz,
                    activo=activo,
                    modified_by = user,
                )
                broker.contactos.set(contactos)  # Relación muchos a muchos
                messages.success(request, 'Se creó un nuevo elemento de forma exitosas.')
            except Exception as e:
                messages.error(request, f'Error: No se pudo crear el elemento. Detalles: {str(e)}')
 
        else:
            try:
                PendingChange.objects.create(
                    model_name='broker',
                    object_id=None,  # No existe aún
                    changes=changes,
                    submitted_by=user,
                    action_type='create',
                )
                messages.success(request, 'Solicitud de creación enviada para aprobación.')
            except Exception as e:
                messages.error(request, f'Error: No se pudo enviar la solicitud. Detalles: {str(e)}')

        return HttpResponseRedirect(request.path_info)
    
    def generar_excel_brokers(self, brokers):
        # Crear un nuevo archivo Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Brokers"

        # Estilos
        font_header = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        relleno_header = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        bordes = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )

        # Encabezados
        encabezados = ["Nombre", "País", "Oficina", "Web", "Matriz", "Activo", "Contactos"]
        for col_num, header in enumerate(encabezados, start=1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = font_header
            cell.fill = relleno_header
            cell.border = bordes

        # Datos
        for row_num, broker in enumerate(brokers, start=2):
            # Columna 1: Nombre
            cell_nombre = sheet.cell(row=row_num, column=1, value=broker.nombre)
            cell_nombre.border = bordes

            # Columna 2: País
            cell_pais = sheet.cell(row=row_num, column=2, value=broker.pais.nombre)
            cell_pais.border = bordes

            # Columna 3: Oficina
            cell_oficina = sheet.cell(row=row_num, column=3, value=broker.domicilio_oficina)
            cell_oficina.border = bordes

            # Columna 4: Web
            cell_web = sheet.cell(row=row_num, column=4, value=broker.url_web)
            cell_web.border = bordes

            # Columna 5: Matriz
            cell_matriz = sheet.cell(row=row_num, column=5, value=broker.matriz.nombre)
            cell_matriz.border = bordes

            # Columna 6: Activo
            cell_activo = sheet.cell(row=row_num, column=6, value="SI" if broker.activo else "NO")
            cell_activo.border = bordes

            # Columna 7: Contactos
            # Obtener nombres de contactos separados por comas
            contactos_nombres = ", ".join(contacto.nombre for contacto in broker.contactos.all())
            sheet.cell(row=row_num, column=7, value=contactos_nombres).border = bordes

        # Crear la respuesta HTTP
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="brokers.xlsx"'
        workbook.save(response)
        workbook.close()
        return response


@method_decorator(login_required, name='dispatch')
class EditarBrokerView(View):
    def post(self, request, broker_id):
        broker = get_object_or_404(Broker, id=broker_id)
        nombre = request.POST.get('editar_nombre')
        logo = request.FILES.get('editar_logo')  # Obtener el logo del archivo subido
        pais_id = request.POST.get('editar_pais')
        domicilio = request.POST.get('editar_domicilio')
        web = request.POST.get('editar_web')
        matriz_id = request.POST.get('editar_matriz')
        activo = 'editar_activo' in request.POST
        contactos_ids = request.POST.getlist('editar_contacto') 
        
        # Obtener objetos relacionados
        pais = get_object_or_404(Pais, id=pais_id)
        matriz = get_object_or_404(Matriz, id=matriz_id)
        contactos = get_list_or_404(Contacto, id__in=contactos_ids)
        user = request.user

        # Si el usuario es superuser, aplicar los cambios directamente
        if user.is_superuser:
            if broker.nombre != nombre:
                broker.nombre = nombre
            if logo:  # Solo actualizar si se sube un nuevo logo
                broker.logo = logo
            if broker.pais_id != int(pais_id):
                broker.pais_id = pais_id
            if broker.domicilio_oficina != domicilio:
                broker.domicilio_oficina = domicilio
            if broker.url_web != web:
                broker.url_web = web
            if broker.matriz_id != int(matriz_id):
                broker.matriz_id = matriz_id
            if broker.activo != activo:
                broker.activo = activo

            broker.contactos.set(contactos)  # Actualiza los contactos asociados al broker
            broker.modified_by = user  # Registrar quién realizó el cambio
            broker.save(track_changes=False)  # Evitar registrar PendingChange
            messages.success(request, 'Los cambios se han aplicado directamente.')
            return redirect('brokers_admin')
        else:
            # Para usuarios no superuser, registrar los cambios pendientes
            changes = {}
            if broker.nombre != nombre:
                changes['nombre'] = {'old': broker.nombre, 'new': nombre}
                broker.nombre = nombre 

            if logo:  # Solo registrar cambio si se sube un logo
                changes['logo'] = {
                    'old': broker.logo.url if broker.logo else None,
                    'new': logo.name
                }
                broker.logo = logo

            if broker.pais_id != int(pais_id):
                changes['pais_id'] = {'old': broker.pais_id, 'new': int(pais_id)}
                broker.pais_id = pais_id

            if broker.domicilio_oficina != domicilio:
                changes['domicilio'] = {'old': broker.domicilio_oficina, 'new': domicilio}
                broker.domicilio_oficina = domicilio 

            if broker.url_web != web:
                changes['web'] = {'old': broker.url_web, 'new': web}
                broker.url_web = web 

            if broker.matriz_id != int(matriz_id):
                changes['matriz_id'] = {'old': broker.matriz_id, 'new': int(matriz_id)}
                broker.matriz_id = matriz_id

            if broker.activo != activo:
                changes['activo'] = {'old': broker.activo, 'new': activo}
                broker.activo = activo

            # Comparar contactos, actualizando la relación ManyToMany
            if set(broker.contactos.all()) != set(contactos):
                changes['contactos'] = {
                    'old': [contacto.id for contacto in broker.contactos.all()],
                    'new': [contacto.id for contacto in contactos]
                }

            if changes:
                try:
                    PendingChange.objects.create(
                        model_name='broker',    
                        object_id=broker.id,
                        changes=changes,
                        submitted_by=user,
                        action_type='edit'
                    )
                    messages.success(request, 'El cambio ha sido registrado para revisión.')
                except Exception as e:
                    messages.error(request, f'Error al registrar el cambio: {str(e)}')
            else:
                messages.info(request, 'No se detectaron cambios.')

        return redirect('brokers_admin')
    
@method_decorator(login_required, name='dispatch')
class EliminarBrokerView(View):
    def post(self, request, broker_id):
        broker = get_object_or_404(Broker, id=broker_id)
        user = request.user

        if user.is_superuser:
            broker.delete()
            messages.success(request, 'El elemento ha sido eliminado directamente.')
            return redirect('brokers_admin')

        changes = {
            'activo': {'old': broker.activo, 'new': False},  # Simula desactivación
        }
        try:
            PendingChange.objects.create(
                model_name='broker',
                object_id=broker.id,
                changes=changes,
                submitted_by=user,
                action_type='delete'  # Indica que es una solicitud de eliminación
            )
            messages.success(request, 'El cambio de eliminación se ha registrado para revisión.')
        except Exception as e:
            messages.error(request, f'Error al registrar el cambio: {str(e)}')

        return redirect('brokers_admin')

# RESUMEN
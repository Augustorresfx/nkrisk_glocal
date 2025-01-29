from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponseRedirect
from django.views.generic import View
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.shortcuts import redirect

# Importe Modelos
from ..models import Pais, Matriz, Broker, Aseguradora, PendingChange

# REPORTE
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from django.http import HttpResponse

# ADMINISTRACION
@method_decorator(login_required, name='dispatch')
class MatrizView(View):
    def get(self, request, *args, **kwargs):
        
        # Obtener filtros desde los parámetros GET
        nombre_filtro = request.GET.get('nombre', '').strip()
        pais_filtro = request.GET.get('pais', '').strip()
        exportar = request.GET.get('exportar', None)  # Verificar si se solicitó exportar
        # Filtrar las matrices con base en los filtros
        matrices = Matriz.objects.all()

        if nombre_filtro:
            matrices = matrices.filter(nombre__icontains=nombre_filtro)

        if pais_filtro:
            matrices = matrices.filter(pais__id=pais_filtro)

        # Paginación
        matrices_paginados = Paginator(matrices, 15)
        page_number = request.GET.get("page")
        filter_pages = matrices_paginados.get_page(page_number)

        # Obtener la lista de países
        paises = Pais.objects.all()

        # Si se solicita exportar, generar el archivo Excel
        if exportar:
            return self.generar_excel_matrices(matrices)

        context = {
            'matrices': matrices,
            'pages': filter_pages,
            'paises': paises,
            'filtros': {
                'nombre': nombre_filtro,
                'pais': pais_filtro,
            }
        }
        return render(request, 'administracion/matrices_admin.html', context)

    def post(self, request, *args, **kwargs):
        # Obtener datos del formulario
        nombre = request.POST.get('nuevo_nombre')
        pais_id = request.POST.get('nuevo_pais')
        activo = request.POST.get('nuevo_activo')
        if activo == "on":
            activo = True
        else:
            activo = False
        pais = get_object_or_404(Pais, id=pais_id)
        user = request.user

        # Crear una solicitud de creación pendiente
        changes = {
            'nombre': {'new': nombre},
            'pais_id': {'new': pais.id},
            'activo': {'new': activo},
        }
        if user.is_superuser:
            try:
                Matriz.objects.create(
                nombre = nombre,
                pais=pais,
                activo=activo,
                modified_by = user,
                )
                messages.success(request, 'Se creó un nuevo elemento de forma exitosa.')
            except Exception as e:
                messages.error(request, f'Error: No se pudo crear el elemento. Detalles: {str(e)}')
 
        else:
            try:
                PendingChange.objects.create(
                    model_name='matriz',
                    object_id=None,  # No existe aún
                    changes=changes,
                    submitted_by=user,
                    action_type='create',
                )
                messages.success(request, 'Solicitud de creación enviada para aprobación.')
            except Exception as e:
                messages.error(request, f'Error: No se pudo enviar la solicitud. Detalles: {str(e)}')

        return HttpResponseRedirect(request.path_info)
    
    def generar_excel_matrices(self, matrices):
        # Crear un nuevo archivo Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Matrices"

        # Estilos
        font_header = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        relleno_header = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        bordes = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )

        # Encabezados
        encabezados = ["Nombre", "País", "Activo"]
        for col_num, header in enumerate(encabezados, start=1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = font_header
            cell.fill = relleno_header
            cell.border = bordes

        # Datos
        for row_num, matriz in enumerate(matrices, start=2):
            # Columna 1: Nombre
            cell_nombre = sheet.cell(row=row_num, column=1, value=matriz.nombre)
            cell_nombre.border = bordes

            # Columna 2: País
            cell_pais = sheet.cell(row=row_num, column=2, value=matriz.pais.nombre)
            cell_pais.border = bordes

            # Columna 3: Activo
            cell_activo = sheet.cell(row=row_num, column=3, value="SI" if matriz.activo else "NO")
            cell_activo.border = bordes

        # Crear la respuesta HTTP
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="matrices.xlsx"'
        workbook.save(response)
        workbook.close()
        return response

    
@method_decorator(login_required, name='dispatch')
class EditarMatrizView(View):
    def post(self, request, matriz_id):
        matriz = get_object_or_404(Matriz, id=matriz_id)
        nombre = request.POST.get('editar_nombre')
        pais_id = request.POST.get('editar_pais')
        activo = 'editar_activo' in request.POST
        user = request.user

        # Si el usuario es superuser, aplicar los cambios directamente
        if user.is_superuser:
            if matriz.nombre != nombre:
                matriz.nombre = nombre
            if matriz.pais_id != int(pais_id):
                matriz.pais_id = pais_id
            if matriz.activo != activo:
                matriz.activo = activo

            matriz.modified_by = user  # Registrar quién realizó el cambio
            matriz.save(track_changes=False)  # Evitar registrar PendingChange
            messages.success(request, 'Los cambios se han aplicado directamente.')
            return redirect('matrices_admin')
        else:
            # Para usuarios no superuser, registrar los cambios pendientes
            changes = {}
            if matriz.nombre != nombre:
                changes['nombre'] = {'old': matriz.nombre, 'new': nombre}
                matriz.nombre = nombre

            if matriz.pais_id != int(pais_id):
                changes['pais_id'] = {'old': matriz.pais_id, 'new': int(pais_id)}
                matriz.pais_id = pais_id

            if matriz.activo != activo:
                changes['activo'] = {'old': matriz.activo, 'new': activo}
                matriz.activo = activo

            if changes:
                try:
                    PendingChange.objects.create(
                        model_name='matriz',
                        object_id=matriz.id,
                        changes=changes,
                        submitted_by=user,
                        action_type='edit'
                    )
                    messages.success(request, 'El cambio ha sido registrado para revisión.')
                except Exception as e:
                    messages.error(request, f'Error al registrar el cambio: {str(e)}')
            else:
                messages.info(request, 'No se detectaron cambios.')

        return redirect('matrices_admin')

@method_decorator(login_required, name='dispatch')
class EliminarMatrizView(View):
    def post(self, request, matriz_id):
        matriz = get_object_or_404(Matriz, id=matriz_id)
        user = request.user

        if user.is_superuser:
            matriz.delete()
            messages.success(request, 'El elemento ha sido eliminado directamente.')
            return redirect('matrices_admin')

        changes = {
            'activo': {'old': matriz.activo, 'new': False},  # Simula desactivación
        }
        try:
            PendingChange.objects.create(
                model_name='matriz',
                object_id=matriz.id,
                changes=changes,
                submitted_by=user,
                action_type='delete'  # Indica que es una solicitud de eliminación
            )
            messages.success(request, 'El cambio de eliminación se ha registrado para revisión.')
        except Exception as e:
            messages.error(request, f'Error al registrar el cambio: {str(e)}')

        return redirect('matrices_admin')
    
# RESUMEN


from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponseRedirect
from django.views.generic import View
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.shortcuts import redirect
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.shortcuts import get_list_or_404

# Importe Modelos
from ..models import Pais, Archivo, PendingChange, Matriz, Contacto, Broker, Categoria

# REPORTE
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from django.http import HttpResponse

# ADMINISTRACION
@method_decorator(login_required, name='dispatch')
class ArchivoView(View):
    def get(self, request, *args, **kwargs):
        
        # Obtener filtros desde los parámetros GET
        nombre_filtro = request.GET.get('nombre', '').strip()
        pais_filtro = request.GET.get('pais', '').strip()
        categoria_filtro = request.GET.get('categoria', '').strip()
        exportar = request.GET.get('exportar', None)
        
        # Filtrar los archivos con base en los filtros 
        archivos = Archivo.objects.all()

        if nombre_filtro:
            archivos = archivos.filter(nombre__icontains=nombre_filtro)

        if pais_filtro:
            archivos = archivos.filter(pais__id=pais_filtro)
            
        if categoria_filtro:
            archivos = archivos.filter(categoria__id=categoria_filtro)
    
        # Paginación
        archivos_paginados = Paginator(archivos, 30)
        page_number = request.GET.get("page")
        filter_pages = archivos_paginados.get_page(page_number)

        # Si se solicita exportar, generar el archivo Excel
        if exportar:
            return self.generar_excel_archivos(archivos)

        # Obtener la lista de países
        paises = Pais.objects.all()

        matrices = Matriz.objects.all()

        contactos = Contacto.objects.all()
        
        categorias = Categoria.objects.all()
        
        brokers = Broker.objects.all()

        context = {
            'archivos': archivos,
            'pages': filter_pages,
            'paises': paises,
            'matrices': matrices,
            'contactos': contactos,
            'categorias': categorias,
            'brokers': brokers,
            'filtros': {
                'nombre': nombre_filtro,
                'pais': pais_filtro,
            }
        }
        return render(request, 'administracion/archivos_admin.html', context)
    
    def post(self, request, *args, **kwargs):
        # Obtener datos del formulario
        pais_id = request.POST.get('nuevo_pais')
        categoria_id = request.POST.get('nuevo_categoria')
        nombre = request.POST.get('nuevo_nombre')
        broker_id = request.POST.get('nuevo_broker')
        activo = request.POST.get('nuevo_activo')
        archivo_ingresado = request.FILES.get('nuevo_archivo')

        if activo == "on":
            activo = True
        else:
            activo = False
        # Obtener objetos relacionados
        pais = get_object_or_404(Pais, id=pais_id)
        broker = get_object_or_404(Broker, id=broker_id)
        categoria = get_object_or_404(Categoria, id=categoria_id)

        user = request.user
        
        # Crear una solicitud de creación pendiente
        changes = {
            'nombre': {'new': nombre},
            'archivo': {'new': archivo_ingresado.name if archivo_ingresado else None},  # Guardar el nombre del archivo

            'pais_id': {'new': pais.id},
            'categoria_id': {'new': categoria.id},
            'broker_id': {'new': broker.id},
            'usuario_id': {'new': user.id},
            'activo': {'new': activo},
        }

        if user.is_superuser:
            try:
                # Crear el objeto archivo directamente si el usuario es superusuario
                archivo = Archivo.objects.create(
                    pais=pais,
                    nombre=nombre,
                    categoria=categoria,
                    broker=broker,
                    usuario=user,
                    activo=activo,
                    archivo=archivo_ingresado,
                    modified_by = user,
                )

                messages.success(request, 'Se creó un nuevo elemento de forma exitosas.')
            except Exception as e:
                messages.error(request, f'Error: No se pudo crear el elemento. Detalles: {str(e)}')
 
        else:
            try:
                PendingChange.objects.create(
                    model_name='archivo',
                    object_id=None,  # No existe aún
                    changes=changes,
                    submitted_by=user,
                    action_type='create',
                )
                messages.success(request, 'Solicitud de creación enviada para aprobación.')
            except Exception as e:
                messages.error(request, f'Error: No se pudo enviar la solicitud. Detalles: {str(e)}')

        return HttpResponseRedirect(request.path_info)
    
    def generar_excel_archivos(self, archivos):
        # Crear un nuevo archivo Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "archivos"

        # Estilos
        font_header = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        relleno_header = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        bordes = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )

        # Encabezados
        encabezados = ["Nombre", "País", "Oficina", "Web", "Matriz", "Activo", "Contactos"]
        for col_num, header in enumerate(encabezados, start=1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = font_header
            cell.fill = relleno_header
            cell.border = bordes

        # Datos
        for row_num, archivo in enumerate(archivos, start=2):
            # Columna 1: Nombre
            cell_nombre = sheet.cell(row=row_num, column=1, value=archivo.nombre)
            cell_nombre.border = bordes

            # Columna 2: País
            cell_pais = sheet.cell(row=row_num, column=2, value=archivo.pais.nombre)
            cell_pais.border = bordes

            # Columna 3: Oficina
            cell_oficina = sheet.cell(row=row_num, column=3, value=archivo.domicilio_oficina)
            cell_oficina.border = bordes

            # Columna 4: Web
            cell_web = sheet.cell(row=row_num, column=4, value=archivo.url_web)
            cell_web.border = bordes

            # Columna 5: Matriz
            cell_matriz = sheet.cell(row=row_num, column=5, value=archivo.matriz.nombre)
            cell_matriz.border = bordes

            # Columna 6: Activo
            cell_activo = sheet.cell(row=row_num, column=6, value="SI" if archivo.activo else "NO")
            cell_activo.border = bordes

            # Columna 7: Contactos
            # Obtener nombres de contactos separados por comas
            contactos_nombres = ", ".join(contacto.nombre for contacto in archivo.contactos.all())
            sheet.cell(row=row_num, column=7, value=contactos_nombres).border = bordes

        # Crear la respuesta HTTP
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="archivos.xlsx"'
        workbook.save(response)
        workbook.close()
        return response


@method_decorator(login_required, name='dispatch')
class EditarArchivoView(View):
    def post(self, request, archivo_id):
        pais_id = request.POST.get('editar_pais')
        categoria_id = request.POST.get('editar_categoria')
        nombre = request.POST.get('editar_nombre')
        broker_id = request.POST.get('editar_broker')
        activo = 'editar_activo' in request.POST
        
        # Obtener objetos relacionados
        archivo = get_object_or_404(Archivo, id=archivo_id)
        pais = get_object_or_404(Pais, id=pais_id)
        categoria = get_object_or_404(Categoria, id=categoria_id)
        broker = get_object_or_404(Broker, id=broker_id)
        user = request.user

        # Si el usuario es superuser, aplicar los cambios directamente
        if user.is_superuser:
            if archivo.nombre != nombre:
                archivo.nombre = nombre

            if archivo.pais_id != int(pais_id):
                archivo.pais_id = pais_id

            if archivo.broker_id != int(broker_id):
                archivo.broker_id = broker_id
            if archivo.categoria_id != int(categoria_id):
                archivo.categoria_id = categoria_id
            if archivo.activo != activo:
                archivo.activo = activo

            archivo.modified_by = user  # Registrar quién realizó el cambio
            archivo.save(track_changes=False)  # Evitar registrar PendingChange
            messages.success(request, 'Los cambios se han aplicado directamente.')
            return redirect('archivos_admin')
        else:
            # Para usuarios no superuser, registrar los cambios pendientes
            changes = {}
            if archivo.nombre != nombre:
                changes['nombre'] = {'old': archivo.nombre, 'new': nombre}
                archivo.nombre = nombre 

            if archivo.pais_id != int(pais_id):
                changes['pais_id'] = {'old': archivo.pais_id, 'new': int(pais_id)}
                archivo.pais_id = pais_id

            if archivo.broker_id != int(broker_id):
                changes['broker_id'] = {'old': archivo.broker_id, 'new': int(broker_id)}
                archivo.broker_id = broker_id
                
                
            if archivo.categoria_id != int(broker_id):
                changes['categoria_id'] = {'old': archivo.categoria_id, 'new': int(categoria_id)}
                archivo.categoria_id = categoria_id


            if archivo.activo != activo:
                changes['activo'] = {'old': archivo.activo, 'new': activo}
                archivo.activo = activo

            if changes:
                try:
                    PendingChange.objects.create(
                        model_name='archivo',    
                        object_id=archivo.id,
                        changes=changes,
                        submitted_by=user,
                        action_type='edit'
                    )
                    messages.success(request, 'El cambio ha sido registrado para revisión.')
                except Exception as e:
                    messages.error(request, f'Error al registrar el cambio: {str(e)}')
            else:
                messages.info(request, 'No se detectaron cambios.')

        return redirect('archivos_admin')
    
@method_decorator(login_required, name='dispatch')
class EliminarArchivoView(View):
    def post(self, request, archivo_id):
        archivo = get_object_or_404(Archivo, id=archivo_id)
        user = request.user

        if user.is_superuser:
            archivo.delete()
            messages.success(request, 'El elemento ha sido eliminado directamente.')
            return redirect('archivos_admin')

        changes = {
            'activo': {'old': archivo.activo, 'new': False},  # Simula desactivación
        }
        try:
            PendingChange.objects.create(
                model_name='archivo',
                object_id=archivo.id,
                changes=changes,
                submitted_by=user,
                action_type='delete'  # Indica que es una solicitud de eliminación
            )
            messages.success(request, 'El cambio de eliminación se ha registrado para revisión.')
        except Exception as e:
            messages.error(request, f'Error al registrar el cambio: {str(e)}')

        return redirect('archivos_admin')

# RESUMEN
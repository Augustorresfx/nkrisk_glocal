from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponseRedirect
from django.views.generic import View
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.shortcuts import redirect
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.shortcuts import get_list_or_404

# Importe Modelos
from ..models import Pais, Broker, Aseguradora, PendingChange, Matriz, Contacto

# REPORTE
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from django.http import HttpResponse

# ADMINISTRACION
@method_decorator(login_required, name='dispatch')
class AseguradoraView(View):
    def get(self, request, *args, **kwargs):
        
        # Obtener filtros desde los parámetros GET
        nombre_filtro = request.GET.get('nombre', '').strip()
        pais_filtro = request.GET.get('pais', '').strip()

        # Filtrar con base en los filtros
        aseguradoras = Aseguradora.objects.all()

        if nombre_filtro:
            aseguradoras = aseguradoras.filter(nombre__icontains=nombre_filtro)

        if pais_filtro:
            aseguradoras = aseguradoras.filter(pais__id=pais_filtro)

        exportar = request.GET.get('exportar', None)

        # Si se solicita exportar, generar el archivo Excel
        if exportar:
            return self.generar_excel_aseguradoras(aseguradoras)

        # Paginación
        aseguradoras_paginados = Paginator(aseguradoras, 30)
        page_number = request.GET.get("page")
        filter_pages = aseguradoras_paginados.get_page(page_number)

        # Obtener la lista de países
        paises = Pais.objects.all()


        contactos = Contacto.objects.all()

        context = {
            'aseguradoras': aseguradoras,
            'pages': filter_pages,
            'paises': paises,
            'contactos': contactos,
            'filtros': {
                'nombre': nombre_filtro,
                'pais': pais_filtro,
            }
        }
        return render(request, 'administracion/aseguradoras_admin.html', context)
    def post(self, request, *args, **kwargs):
        # Obtener datos del formulario
        nombre = request.POST.get('nuevo_nombre')
        logo = request.FILES.get('nuevo_logo')
        ruc_nit = request.POST.get('nuevo_ruc_nit')
        pais_id = request.POST.get('nuevo_pais')
        activo = request.POST.get('nuevo_activo')
        contactos_ids = request.POST.getlist('nuevo_contacto')
        if activo == "on":
            activo = True
        else:
            activo = False
        # Obtener objetos relacionados
        pais = get_object_or_404(Pais, id=pais_id)
        contactos = get_list_or_404(Contacto, id__in=contactos_ids)
        user = request.user
        # Crear una solicitud de creación pendiente
        changes = {
            'nombre': {'new': nombre},
            'logo': {'new': logo.name if logo else None},  # Guardar el nombre del archivo
            'ruc_nit': {'new': ruc_nit},
            'pais_id': {'new': pais.id},
            'activo': {'new': activo},
            'contactos': {'new': [c.id for c in contactos]},  # IDs de los contactos seleccionados
        }

        if user.is_superuser:
            try:
                # Crear el objeto directamente si el usuario es superusuario
                aseguradora = Aseguradora.objects.create(
                    nombre=nombre,
                    logo=logo,
                    ruc_nit=ruc_nit,
                    pais=pais,
                    activo=activo,
                    modified_by = user,
                )
                aseguradora.contactos.set(contactos)  # Relación muchos a muchos
                messages.success(request, 'Se creó un nuevo elemento de forma exitosas.')
            except Exception as e:
                messages.error(request, f'Error: No se pudo crear el elemento. Detalles: {str(e)}')
 
        else:
            try:
                PendingChange.objects.create(
                    model_name='aseguradora',
                    object_id=None,  # No existe aún
                    changes=changes,
                    submitted_by=user,
                    action_type='create',
                )
                messages.success(request, 'Solicitud de creación enviada para aprobación.')
            except Exception as e:
                messages.error(request, f'Error: No se pudo enviar la solicitud. Detalles: {str(e)}')

        return HttpResponseRedirect(request.path_info)
    
    def generar_excel_aseguradoras(self, aseguradoras):
        # Crear un nuevo archivo Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "aseguradoras"

        # Estilos
        font_header = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        relleno_header = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        bordes = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )

        # Encabezados
        encabezados = ["Nombre", "País", "RUC / NIT", "Activo", "Contactos"]
        for col_num, header in enumerate(encabezados, start=1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = font_header
            cell.fill = relleno_header
            cell.border = bordes

        # Datos
        for row_num, aseguradora in enumerate(aseguradoras, start=2):
            # Columna 1: Nombre
            cell_nombre = sheet.cell(row=row_num, column=1, value=aseguradora.nombre)
            cell_nombre.border = bordes

            # Columna 2: País
            cell_pais = sheet.cell(row=row_num, column=2, value=aseguradora.pais.nombre)
            cell_pais.border = bordes

            # Columna 3: Ruc / nit
            cell_oficina = sheet.cell(row=row_num, column=3, value=aseguradora.ruc_nit)
            cell_oficina.border = bordes

            # Columna 4: Activo
            cell_activo = sheet.cell(row=row_num, column=4, value="SI" if aseguradora.activo else "NO")
            cell_activo.border = bordes

            # Columna 5: Contactos
            # Obtener nombres de contactos separados por comas
            contactos_nombres = ", ".join(contacto.nombre for contacto in aseguradora.contactos.all())
            sheet.cell(row=row_num, column=5, value=contactos_nombres).border = bordes

        # Crear la respuesta HTTP
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="aseguradoras.xlsx"'
        workbook.save(response)
        workbook.close()
        return response

@method_decorator(login_required, name='dispatch')
class EditarAseguradoraView(View):
    def post(self, request, aseguradora_id):
        aseguradora = get_object_or_404(Aseguradora, id=aseguradora_id)
        nombre = request.POST.get('editar_nombre')
        logo = request.FILES.get('editar_logo')  # Obtener el logo del archivo subido
        pais_id = request.POST.get('editar_pais')
        ruc_nit = request.POST.get('editar_ruc_nit')
        activo = 'editar_activo' in request.POST
        contactos_ids = request.POST.getlist('editar_contacto') 
        
        # Obtener objetos relacionados
        pais = get_object_or_404(Pais, id=pais_id)
        contactos = get_list_or_404(Contacto, id__in=contactos_ids)
        user = request.user

        # Si el usuario es superuser, aplicar los cambios directamente
        if user.is_superuser:
            if aseguradora.nombre != nombre:
                aseguradora.nombre = nombre
            if logo:  # Solo actualizar si se sube un nuevo logo
                aseguradora.logo = logo
            if aseguradora.pais_id != int(pais_id):
                aseguradora.pais_id = pais_id
            if aseguradora.ruc_nit != ruc_nit:
                aseguradora.ruc_nit = ruc_nit

            if aseguradora.activo != activo:
                aseguradora.activo = activo

            aseguradora.contactos.set(contactos)  # Actualiza los contactos asociados
            aseguradora.modified_by = user  # Registrar quién realizó el cambio
            aseguradora.save(track_changes=False)  # Evitar registrar PendingChange
            messages.success(request, 'Los cambios se han aplicado directamente.')
            return redirect('aseguradoras_admin')
        else:
            # Para usuarios no superuser, registrar los cambios pendientes
            changes = {}
            if aseguradora.nombre != nombre:
                changes['nombre'] = {'old': aseguradora.nombre, 'new': nombre}
                aseguradora.nombre = nombre 

            if logo:  # Solo registrar cambio si se sube un logo
                changes['logo'] = {
                    'old': aseguradora.logo.url if aseguradora.logo else None,
                    'new': logo.name
                }
                aseguradora.logo = logo

            if aseguradora.pais_id != int(pais_id):
                changes['pais_id'] = {'old': aseguradora.pais_id, 'new': int(pais_id)}
                aseguradora.pais_id = pais_id

            if aseguradora.ruc_nit != ruc_nit:
                changes['ruc_nit'] = {'old': aseguradora.ruc_nit, 'new': ruc_nit}
                aseguradora.ruc_nit = ruc_nit 

            if aseguradora.activo != activo:
                changes['activo'] = {'old': aseguradora.activo, 'new': activo}
                aseguradora.activo = activo

            # Comparar contactos, actualizando la relación ManyToMany
            if set(aseguradora.contactos.all()) != set(contactos):
                changes['contactos'] = {
                    'old': [contacto.id for contacto in aseguradora.contactos.all()],
                    'new': [contacto.id for contacto in contactos]
                }

            if changes:
                try:
                    PendingChange.objects.create(
                        model_name='aseguradora',    
                        object_id=aseguradora.id,
                        changes=changes,
                        submitted_by=user,
                        action_type='edit'
                    )
                    messages.success(request, 'El cambio ha sido registrado para revisión.')
                except Exception as e:
                    messages.error(request, f'Error al registrar el cambio: {str(e)}')
            else:
                messages.info(request, 'No se detectaron cambios.')

        return redirect('aseguradoras_admin')
    
@method_decorator(login_required, name='dispatch')
class EliminarAseguradoraView(View):
    def post(self, request, aseguradora_id):
        aseguradora = get_object_or_404(Aseguradora, id=aseguradora_id)
        user = request.user

        if user.is_superuser:
            aseguradora.delete()
            messages.success(request, 'El elemento ha sido eliminado directamente.')
            return redirect('aseguradoras_admin')

        changes = {
            'activo': {'old': aseguradora.activo, 'new': False},  # Simula desactivación
        }
        try:
            PendingChange.objects.create(
                model_name='aseguradora',
                object_id=aseguradora.id,
                changes=changes,
                submitted_by=user,
                action_type='delete'  # Indica que es una solicitud de eliminación
            )
            messages.success(request, 'El cambio de eliminación se ha registrado para revisión.')
        except Exception as e:
            messages.error(request, f'Error al registrar el cambio: {str(e)}')

        return redirect('aseguradoras_admin')

# RESUMEN
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponseRedirect
from django.views.generic import View
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.shortcuts import redirect
from django.conf import settings
from django.contrib.auth import get_user_model
# Importe Modelos
from ..models import User, Contacto, PendingChange

# REPORTE
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from django.http import HttpResponse

# ADMINISTRACION
@method_decorator(login_required, name='dispatch')
class ContactoView(View):
    def get(self, request, *args, **kwargs):
        
        # Obtener filtros desde los parámetros GET
        nombre_filtro = request.GET.get('nombre', '').strip()
        email_filtro = request.GET.get('email', '').strip()
        telefono_filtro = request.GET.get('telefono', '').strip()
        cargo_filtro = request.GET.get('cargo', '').strip()

        # Filtrar los contactos con base en los filtros
        contactos = Contacto.objects.all()
        if nombre_filtro:
            contactos = contactos.filter(nombre__icontains=nombre_filtro)

        if email_filtro:
            contactos = contactos.filter(email__icontains=email_filtro)

        if telefono_filtro:
            contactos = contactos.filter(telefono__icontains=telefono_filtro)

        if cargo_filtro:
            contactos = contactos.filter(cargo__icontains=cargo_filtro)
        # Paginación
        contactos_paginados = Paginator(contactos, 30)
        page_number = request.GET.get("page")
        filter_pages = contactos_paginados.get_page(page_number)
        
        exportar = request.GET.get('exportar', None)
        # Si se solicita exportar, generar el archivo Excel
        if exportar:
            return self.generar_excel_contactos(contactos)


        # Obtener el modelo del usuario personalizado
        user_custom = get_user_model()
        print(user_custom)
        # Obtener la lista de usuarios
        usuarios = user_custom.objects.all()
        
        context = {
            'contactos': contactos,
            'pages': filter_pages,
            'usuarios': usuarios,
            'filtros': {
                'nombre': nombre_filtro,
                'email': email_filtro,
                'telefono': telefono_filtro,
                'cargo': cargo_filtro,
            }
        }
        return render(request, 'administracion/contactos_admin.html', context)

    def post(self, request, *args, **kwargs):
        # Obtener datos del formulario
        nombre = request.POST.get('nuevo_nombre')
        email = request.POST.get('nuevo_email')
        telefono = request.POST.get('nuevo_telefono')
        cargo = request.POST.get('nuevo_cargo')
        user_id = request.POST.get('nuevo_user')
        user_custom = get_user_model()
        usuario_contacto = get_object_or_404(user_custom, id=user_id)
       
        user = get_user_model().objects.get(id=request.user.id)
        print(user)
        # Crear una solicitud de creación pendiente
        changes = {
            'nombre': {'new': nombre},
            'email': {'new': email},
            'telefono': {'new': telefono},
            'cargo': {'new': cargo},
            'user_id': {'new': usuario_contacto.id},
        }
        if user.is_superuser:
            try:
                Contacto.objects.create(
                nombre = nombre,
                email=email,
                telefono=telefono,
                cargo=cargo,
                user = usuario_contacto,
                modified_by = user,
                )
                messages.success(request, 'Se creó un nuevo elemento de forma exitosa.')
            except Exception as e:
                messages.error(request, f'Error: No se pudo crear el elemento. Detalles: {str(e)}')
 
        else:
            try:
                PendingChange.objects.create(
                    model_name='contacto',
                    object_id=None,  # No existe aún
                    changes=changes,
                    submitted_by=user,
                    action_type='create',
                )
                messages.success(request, 'Solicitud de creación enviada para aprobación.')
            except Exception as e:
                messages.error(request, f'Error: No se pudo enviar la solicitud. Detalles: {str(e)}')

        return HttpResponseRedirect(request.path_info)
    
    def generar_excel_contactos(self, contactos):
        # Crear un nuevo archivo Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Contactos"

        # Estilos
        font_header = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        relleno_header = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        bordes = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )

        # Encabezados
        encabezados = ["Nombre", "Email", "Teléfono", "Cargo"]
        for col_num, header in enumerate(encabezados, start=1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = font_header
            cell.fill = relleno_header
            cell.border = bordes

        # Datos
        for row_num, contacto in enumerate(contactos, start=2):
            # Columna 1: Nombre
            cell_nombre = sheet.cell(row=row_num, column=1, value=contacto.nombre)
            cell_nombre.border = bordes

            # Columna 2: Email
            cell_pais = sheet.cell(row=row_num, column=2, value=contacto.email)
            cell_pais.border = bordes

            # Columna 3: Telefono
            cell_oficina = sheet.cell(row=row_num, column=3, value=contacto.telefono)
            cell_oficina.border = bordes

            # Columna 4: Cargo
            cell_web = sheet.cell(row=row_num, column=4, value=contacto.cargo)
            cell_web.border = bordes

        # Crear la respuesta HTTP
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="contactos.xlsx"'
        workbook.save(response)
        workbook.close()
        return response

@method_decorator(login_required, name='dispatch')
class EditarContactoView(View):
    def post(self, request, contacto_id):
        contacto = get_object_or_404(Contacto, id=contacto_id)
        nombre = request.POST.get('editar_nombre')
        email = request.POST.get('editar_email')
        telefono = request.POST.get('editar_telefono')
        cargo = request.POST.get('editar_cargo')
        user_id = request.POST.get('editar_usuario')

        # Obtener el usuario que está realizando la modificación
        user = get_user_model().objects.get(id=request.user.id)
        # Obtener el usuario de contacto relacionado por ID
        usuario_contacto = get_user_model().objects.get(id=user_id)
        print("User: ", user, type(user))
        print("Usuario contacto", usuario_contacto, type(usuario_contacto))
        # Si el usuario es superuser, aplicar los cambios directamente
        if user.is_superuser:
            if contacto.nombre != nombre:
                contacto.nombre = nombre
            if contacto.email != email:
                contacto.email = email
            if contacto.telefono != telefono:
                contacto.telefono = telefono
            if contacto.cargo != cargo:
                contacto.cargo = cargo
            if contacto.user != usuario_contacto:  # Cambié la comparación a 'user' en lugar de 'user_id'
                contacto.user = usuario_contacto  # Asignar la instancia del usuario

            contacto.save(track_changes=False)  # Guardar los cambios directamente
            messages.success(request, 'Los cambios se han aplicado directamente.')
            return redirect('contactos_admin')

        # Para usuarios no superuser, registrar los cambios pendientes
        changes = {}
        if contacto.nombre != nombre:
            changes['nombre'] = {'old': contacto.nombre, 'new': nombre}
            contacto.nombre = nombre
            
        if contacto.email != email:
            changes['email'] = {'old': contacto.email, 'new': email}
            contacto.email = email
            
        if contacto.telefono != telefono:
            changes['telefono'] = {'old': contacto.telefono, 'new': telefono}
            contacto.telefono = telefono
            
        if contacto.cargo != cargo:
            changes['cargo'] = {'old': contacto.cargo, 'new': cargo}
            contacto.cargo = cargo

        if contacto.user != usuario_contacto:  # Comprobamos si el usuario es diferente
            changes['user'] = {'old': contacto.user.id, 'new': usuario_contacto.id}
            contacto.user = usuario_contacto

        if changes:
            try:
                PendingChange.objects.create(
                    model_name='contacto',
                    object_id=contacto.id,
                    changes=changes,
                    submitted_by=user,
                    action_type='edit'
                )
                messages.success(request, 'El cambio ha sido registrado para revisión.')
            except Exception as e:
                messages.error(request, f'Error al registrar el cambio: {str(e)}')
        else:
            messages.info(request, 'No se detectaron cambios.')

        return redirect('contactos_admin')


@method_decorator(login_required, name='dispatch')
class EliminarContactoView(View):
    def post(self, request, contacto_id):
        # Obtener el contacto a través de su ID
        contacto = get_object_or_404(Contacto, id=contacto_id)
        user = request.user

        if user.is_superuser:
            # Si el usuario es superusuario, eliminamos el contacto directamente
            contacto.delete()
            messages.success(request, 'El elemento ha sido eliminado correctamente.')
            return redirect('contactos_admin')

        # Si no es superusuario, registramos la solicitud de eliminación sin modificar el modelo Contacto
        try:
            # Crear un cambio pendiente para marcar la eliminación
            PendingChange.objects.create(
                model_name='contacto',
                object_id=contacto.id,
                changes={'deleted': {'old': None, 'new': True}},  # Simulamos una eliminación
                submitted_by=user,
                action_type='delete'  # Acción de eliminación
            )
            messages.success(request, 'El cambio de eliminación se ha registrado para revisión.')
        except Exception as e:
            messages.error(request, f'Error al registrar el cambio: {str(e)}')

        return redirect('contactos_admin')

# RESUMEN